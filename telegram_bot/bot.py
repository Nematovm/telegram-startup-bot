import logging
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardRemove, BotCommand, MenuButtonCommands
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes, ConversationHandler, CallbackQueryHandler
from datetime import datetime
import openpyxl
from openpyxl import Workbook
import os

# Logging sozlash
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)

# Bosqichlar
ISM, FAMILIYA, TELEFON, KANALLAR, MAQSAD, BROADCAST_MESSAGE = range(6)

# ADMIN ID (o'zingizning Telegram ID ingizni kiriting!)
# ID ni olish uchun: @userinfobot ga /start yozing
ADMIN_IDS = [5304449378]  # Bu yerga o'z ID ingizni kiriting!

# Tekshirish kerak bo'lgan kanallar
REQUIRED_CHANNELS = [
    "@muhammadali_journey"
]

# Kanal nomlarini ko'rsatish uchun
CHANNEL_NAMES = {
    "@muhammadali_journey": "Muhammad Ali Journey"
}

# Excel faylga saqlash funksiyasi
def save_to_excel(data):
    filename = 'users_data.xlsx'
    
    # Agar fayl mavjud bo'lmasa, yangi yaratamiz
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Foydalanuvchilar"
        ws.append(['№', 'Ism', 'Familiya', 'Telefon', 'Maqsad', 'Telegram ID', 'Username', 'Sana'])
        wb.save(filename)
    
    # Faylni ochamiz
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    # Yangi qatorni qo'shamiz
    row_num = ws.max_row + 1
    ws.append([
        row_num - 1,
        data.get('ism', ''),
        data.get('familiya', ''),
        data.get('telefon', ''),
        data.get('maqsad', ''),
        data.get('user_id', ''),
        data.get('username', ''),
        datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ])
    
    wb.save(filename)
    return filename

# User allaqachon ro'yxatdan o'tganmi tekshirish
def is_user_registered(user_id):
    filename = 'users_data.xlsx'
    if not os.path.exists(filename):
        return False
    
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    # Telegram ID ustunida qidirish (6-ustun)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[5] == user_id:  # 5-indeks = Telegram ID
            return True
    return False

# Barcha userlarni olish
def get_all_users():
    filename = 'users_data.xlsx'
    if not os.path.exists(filename):
        return []
    
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    users = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        users.append({
            'num': row[0],
            'ism': row[1],
            'familiya': row[2],
            'telefon': row[3],
            'maqsad': row[4],
            'user_id': row[5],
            'username': row[6],
            'sana': row[7]
        })
    return users

# Statistika olish
def get_statistics():
    users = get_all_users()
    if not users:
        return "Hali hech kim ro'yxatdan o'tmagan"
    
    total = len(users)
    today = datetime.now().strftime('%Y-%m-%d')
    today_count = sum(1 for u in users if str(u['sana']).startswith(today))
    
    return f"""
📊 STATISTIKA

👥 Jami ro'yxatdan o'tganlar: {total} ta
📅 Bugun ro'yxatdan o'tganlar: {today_count} ta
⏰ Oxirgi ro'yxat: {users[-1]['sana'] if users else 'Mavjud emas'}
    """

# Admin tekshirish
def is_admin(user_id):
    return user_id in ADMIN_IDS

# Asosiy menyu (ro'yxatdan o'tganlar uchun)
def get_main_menu():
    keyboard = [
        ["ℹ️ Ma'lumot", "📞 Bog'lanish"],
        ["📊 Statistika", "🔔 Yangiliklar"]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

# Admin menyu
def get_admin_menu():
    keyboard = [
        ["📊 Statistika", "👥 Userlar"],
        ["📢 Xabar yuborish", "📁 Excel yuklab olish"],
        ["🔙 Oddiy menyu"]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

# /start buyrug'i
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    user_name = update.effective_user.first_name
    
    # Agar user allaqachon ro'yxatdan o'tgan bo'lsa
    if is_user_registered(user_id):
        # Admin uchun maxsus menyu
        if is_admin(user_id):
            menu = get_admin_menu()
            xabar = f"👋 Xush kelibsiz, Admin {user_name}!\n\n📋 Admin panel ochiq."
        else:
            menu = get_main_menu()
            xabar = f"👋 Xush kelibsiz, {user_name}!\n\n✅ Siz allaqachon ro'yxatdan o'tgansiz!"
        
        await update.message.reply_text(xabar, reply_markup=menu)
        return ConversationHandler.END
    
    # Eski ma'lumotlarni tozalash
    context.user_data.clear()
    
    await update.message.reply_text(
        f"Assalomu alaykum, {user_name}! 👋\n\n"
        "Ro'yxatdan o'tish uchun ismingizni kiriting:",
        reply_markup=ReplyKeyboardRemove()
    )
    return ISM

# Ism qabul qilish
async def ism_qabul(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['ism'] = update.message.text
    await update.message.reply_text(
        f"Yaxshi, {update.message.text}! Endi familiyangizni kiriting:"
    )
    return FAMILIYA

# Familiya qabul qilish
async def familiya_qabul(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['familiya'] = update.message.text
    
    keyboard = [[KeyboardButton("📱 Telefon raqamni yuborish", request_contact=True)]]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True)
    
    await update.message.reply_text(
        "Ajoyib! Endi telefon raqamingizni yuboring:",
        reply_markup=reply_markup
    )
    return TELEFON

# Telefon raqam qabul qilish
async def telefon_qabul(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.contact:
        telefon = update.message.contact.phone_number
    else:
        telefon = update.message.text
    
    context.user_data['telefon'] = telefon
    user_id = update.effective_user.id
    
    # Kanallarga a'zolikni tekshirish
    azo_emas = []
    
    for kanal in REQUIRED_CHANNELS:
        try:
            member = await context.bot.get_chat_member(kanal, user_id)
            if member.status not in ['member', 'administrator', 'creator']:
                azo_emas.append(kanal)
        except Exception as e:
            logging.error(f"Kanal tekshirishda xato {kanal}: {e}")
            azo_emas.append(kanal)
    
    if not azo_emas:
        await update.message.reply_text(
            "Ajoyib! Siz barcha kanallarga a'zosiz ✅\n\n"
            "Endi qatnashish maqsadingizni yozing:",
            reply_markup=ReplyKeyboardRemove()
        )
        return MAQSAD
    else:
        keyboard = []
        for kanal in azo_emas:
            kanal_nomi = CHANNEL_NAMES.get(kanal, kanal)
            if kanal.startswith('@'):
                url = f"https://t.me/{kanal[1:]}"
            else:
                url = f"https://t.me/{REQUIRED_CHANNELS[0][1:]}"
            keyboard.append([InlineKeyboardButton(f"A'zo bo'lish: {kanal_nomi}", url=url)])
        
        keyboard.append([InlineKeyboardButton("✅ Tekshirish", callback_data="check_membership")])
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await update.message.reply_text(
            "❌ Iltimos, avval quyidagi kanallarga a'zo bo'ling:\n\n" + 
            "\n".join([CHANNEL_NAMES.get(k, k) for k in azo_emas]) + 
            "\n\nA'zo bo'lganingizdan keyin 'Tekshirish' tugmasini bosing.",
            reply_markup=reply_markup
        )
        return KANALLAR

# Kanal a'zoligini qayta tekshirish
async def kanallarni_tekshir(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    user_id = update.effective_user.id
    azo_emas = []
    
    for kanal in REQUIRED_CHANNELS:
        try:
            member = await context.bot.get_chat_member(kanal, user_id)
            if member.status not in ['member', 'administrator', 'creator']:
                azo_emas.append(kanal)
        except Exception as e:
            logging.error(f"Kanal tekshirishda xato {kanal}: {e}")
            azo_emas.append(kanal)
    
    if not azo_emas:
        await query.edit_message_text(
            "✅ Ajoyib! Siz barcha kanallarga a'zosiz!\n\n"
            "Endi qatnashish maqsadingizni yozing:"
        )
        return MAQSAD
    else:
        keyboard = []
        for kanal in azo_emas:
            kanal_nomi = CHANNEL_NAMES.get(kanal, kanal)
            if kanal.startswith('@'):
                url = f"https://t.me/{kanal[1:]}"
            else:
                url = f"https://t.me/{REQUIRED_CHANNELS[0][1:]}"
            keyboard.append([InlineKeyboardButton(f"A'zo bo'lish: {kanal_nomi}", url=url)])
        
        keyboard.append([InlineKeyboardButton("✅ Tekshirish", callback_data="check_membership")])
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        try:
            await query.edit_message_text(
                "❌ Hali ham quyidagi kanallarga a'zo emassiz:\n\n" + 
                "\n".join([CHANNEL_NAMES.get(k, k) for k in azo_emas]) + 
                "\n\nIltimos, a'zo bo'ling va qayta tekshiring.",
                reply_markup=reply_markup
            )
        except Exception:
            pass
        return KANALLAR

# Maqsadni qabul qilish va yakunlash
async def maqsad_qabul(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['maqsad'] = update.message.text
    context.user_data['user_id'] = update.effective_user.id
    context.user_data['username'] = update.effective_user.username or 'Username yo\'q'
    
    # Excel ga saqlash
    try:
        filename = save_to_excel(context.user_data)
        save_status = f"\n💾 Ma'lumotlar saqlandi!"
    except Exception as e:
        logging.error(f"Excel ga saqlashda xato: {e}")
        save_status = "\n⚠️ Ma'lumotlarni saqlashda xatolik yuz berdi"
    
    # User ga javob
    malumot = f"""
✅ Ro'yxatdan o'tish muvaffaqiyatli yakunlandi!

📝 Sizning ma'lumotlaringiz:
👤 Ism: {context.user_data.get('ism')}
👤 Familiya: {context.user_data.get('familiya')}
📱 Telefon: {context.user_data.get('telefon')}
🎯 Maqsad: {context.user_data.get('maqsad')}
{save_status}

Rahmat! ✨
    """
    
    await update.message.reply_text(malumot, reply_markup=get_main_menu())
    
    # Adminga xabar yuborish
    admin_notification = f"""
🆕 YANGI RO'YXAT!

👤 Ism: {context.user_data.get('ism')} {context.user_data.get('familiya')}
📱 Telefon: {context.user_data.get('telefon')}
🎯 Maqsad: {context.user_data.get('maqsad')}
🆔 Telegram ID: {context.user_data.get('user_id')}
👤 Username: @{context.user_data.get('username')}
⏰ Vaqt: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
    """
    
    for admin_id in ADMIN_IDS:
        try:
            await context.bot.send_message(chat_id=admin_id, text=admin_notification)
        except Exception as e:
            logging.error(f"Adminga xabar yuborishda xato: {e}")
    
    context.user_data.clear()
    return ConversationHandler.END

# MENU TUGMALARI HANDLERI
async def handle_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    user_id = update.effective_user.id
    
    # Admin tugmalari
    if is_admin(user_id):
        if text == "📊 Statistika":
            await stats_command(update, context)
            return
        elif text == "👥 Userlar":
            await users_command(update, context)
            return
        elif text == "📢 Xabar yuborish":
            await broadcast_command(update, context)
            return BROADCAST_MESSAGE
        elif text == "📁 Excel yuklab olish":
            if os.path.exists('users_data.xlsx'):
                await update.message.reply_document(
                    document=open('users_data.xlsx', 'rb'),
                    filename='users_data.xlsx',
                    caption="📊 Barcha foydalanuvchilar ma'lumotlari"
                )
            else:
                await update.message.reply_text("Excel fayl topilmadi")
            return
        elif text == "🔙 Oddiy menyu":
            await update.message.reply_text(
                "Oddiy menyu:",
                reply_markup=get_main_menu()
            )
            return
    
    # Oddiy user tugmalari
    if text == "ℹ️ Ma'lumot":
        await update.message.reply_text(
            "ℹ️ BIZ HAQIMIZDA\n\n"
            "Bu bot Start-up loyihasi uchun yaratilgan.\n"
            "Bizning kanallarimizga a'zo bo'ling va yangiliklar bilan tanishing!"
        )
    elif text == "📞 Bog'lanish":
        await update.message.reply_text(
            "📞 BOG'LANISH\n\n"
            "Savollaringiz bo'lsa, biz bilan bog'laning:\n"
            "📧 Email: nematovm764@gmail.com\n"
            "📱 Telefon: +998 91 642 81 86"
        )
    elif text == "📊 Statistika":
        users_count = len(get_all_users())
        await update.message.reply_text(
            f"📊 STATISTIKA\n\n"
            f"👥 Jami a'zolar: {users_count} ta"
        )
    elif text == "🔔 Yangiliklar":
        await update.message.reply_text(
            "🔔 YANGILIKLAR\n\n"
            "Yangiliklar tez orada e'lon qilinadi!\n"
            "Bizning kanallarni kuzatib boring."
        )
    elif text == "🔧 Admin Panel" and is_admin(user_id):
        await update.message.reply_text(
            "🔧 Admin Panel:",
            reply_markup=get_admin_menu()
        )

# ADMIN BUYRUQLARI
async def stats_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        await update.message.reply_text("❌ Bu buyruq faqat adminlar uchun!")
        return
    
    stats = get_statistics()
    menu = get_admin_menu() if is_admin(user_id) else None
    await update.message.reply_text(stats, reply_markup=menu)

async def users_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        await update.message.reply_text("❌ Bu buyruq faqat adminlar uchun!")
        return
    
    users = get_all_users()
    
    if not users:
        await update.message.reply_text("Hali hech kim ro'yxatdan o'tmagan", reply_markup=get_admin_menu())
        return
    
    # Oxirgi 10 ta userni ko'rsatish
    last_users = users[-10:]
    message = "👥 OXIRGI 10 TA USER:\n\n"
    
    for user in last_users:
        message += f"#{user['num']} {user['ism']} {user['familiya']}\n"
        message += f"📱 {user['telefon']}\n"
        message += f"🆔 {user['user_id']}\n"
        message += f"📅 {user['sana']}\n\n"
    
    await update.message.reply_text(message, reply_markup=get_admin_menu())
    
    # Excel faylni yuborish
    if os.path.exists('users_data.xlsx'):
        await update.message.reply_document(
            document=open('users_data.xlsx', 'rb'),
            filename='users_data.xlsx',
            caption="📊 To'liq ro'yxat Excel faylda"
        )

async def broadcast_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        await update.message.reply_text("❌ Bu buyruq faqat adminlar uchun!")
        return ConversationHandler.END
    
    await update.message.reply_text(
        "📢 Barcha foydalanuvchilarga yubormoqchi bo'lgan xabaringizni yozing:\n\n"
        "Bekor qilish uchun /cancel ni yozing.",
        reply_markup=ReplyKeyboardRemove()
    )
    return BROADCAST_MESSAGE

async def send_broadcast(update: Update, context: ContextTypes.DEFAULT_TYPE):
    message_text = update.message.text
    users = get_all_users()
    
    if not users:
        await update.message.reply_text("Hali hech kim ro'yxatdan o'tmagan")
        return ConversationHandler.END
    
    success_count = 0
    fail_count = 0
    
    status_msg = await update.message.reply_text(f"📤 Yuborilmoqda... 0/{len(users)}")
    
    for i, user in enumerate(users):
        try:
            await context.bot.send_message(chat_id=user['user_id'], text=message_text)
            success_count += 1
        except Exception as e:
            logging.error(f"User {user['user_id']} ga yuborishda xato: {e}")
            fail_count += 1
        
        # Har 5 ta userdan keyin statusni yangilash
        if (i + 1) % 5 == 0:
            await status_msg.edit_text(f"📤 Yuborilmoqda... {i+1}/{len(users)}")
    
    await status_msg.edit_text(
        f"✅ Xabar yuborish tugadi!\n\n"
        f"✅ Muvaffaqiyatli: {success_count}\n"
        f"❌ Xato: {fail_count}\n"
        f"📊 Jami: {len(users)}"
    )
    
    # Admin menyuga qaytish
    await update.message.reply_text(
        "Admin panel:",
        reply_markup=get_admin_menu()
    )
    
    return ConversationHandler.END

async def bekor(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    
    # Agar ro'yxatdan o'tgan bo'lsa, menyuga qaytarish
    if is_user_registered(user_id):
        menu = get_admin_menu() if is_admin(user_id) else get_main_menu()
        await update.message.reply_text("Jarayon bekor qilindi.", reply_markup=menu)
    else:
        await update.message.reply_text(
            "Jarayon bekor qilindi. /start ni bosing qaytadan boshlash uchun.",
            reply_markup=ReplyKeyboardRemove()
        )
    
    context.user_data.clear()
    return ConversationHandler.END

def main():
    TOKEN = "8306503032:AAFs69tXbgDuZN4ZPVmC3AN3FM9tkR0rtiU"
    
    application = Application.builder().token(TOKEN).build()
    
    # Bot buyruqlarini va Menu tugmasini sozlash
    async def post_init(app: Application):
        # Buyruqlar ro'yxati
        commands = [
            BotCommand("start", "🏠 Botni ishga tushirish"),
            BotCommand("stats", "📊 Statistika (Admin)"),
            BotCommand("users", "👥 Foydalanuvchilar ro'yxati (Admin)"),
            BotCommand("broadcast", "📢 Xabar yuborish (Admin)"),
            BotCommand("cancel", "❌ Jarayonni bekor qilish"),
        ]
        await app.bot.set_my_commands(commands)
        
        # Menu tugmasini yoqish
        await app.bot.set_chat_menu_button(menu_button=MenuButtonCommands())
    
    application.post_init = post_init
    
    # Asosiy suhbat
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            ISM: [MessageHandler(filters.TEXT & ~filters.COMMAND, ism_qabul)],
            FAMILIYA: [MessageHandler(filters.TEXT & ~filters.COMMAND, familiya_qabul)],
            TELEFON: [MessageHandler((filters.CONTACT | filters.TEXT) & ~filters.COMMAND, telefon_qabul)],
            KANALLAR: [CallbackQueryHandler(kanallarni_tekshir, pattern="check_membership")],
            MAQSAD: [MessageHandler(filters.TEXT & ~filters.COMMAND, maqsad_qabul)],
        },
        fallbacks=[CommandHandler('cancel', bekor)],
        allow_reentry=True
    )
    
    # Broadcast suhbati
    broadcast_handler = ConversationHandler(
        entry_points=[
            CommandHandler('broadcast', broadcast_command),
            MessageHandler(filters.Regex('^📢 Xabar yuborish$'), broadcast_command)
        ],
        states={
            BROADCAST_MESSAGE: [MessageHandler(filters.TEXT & ~filters.COMMAND, send_broadcast)],
        },
        fallbacks=[CommandHandler('cancel', bekor)]
    )
    
    application.add_handler(conv_handler)
    application.add_handler(broadcast_handler)
    
    # Admin buyruqlari
    application.add_handler(CommandHandler('stats', stats_command))
    application.add_handler(CommandHandler('users', users_command))
    
    # Menu tugmalari handleri
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_menu))
    
    print("🚀 Bot ishga tushdi!")
    print("📋 Menu tugmasi qo'shildi! (Rasmingiz yonida 'Menu' ko'rinadi)")
    print("\n📊 Admin buyruqlari:")
    print("   /stats - Statistika")
    print("   /users - Barcha userlar")
    print("   /broadcast - Hammaga xabar yuborish")
    print("\n💡 Eslatma: ADMIN_IDS da o'z ID ingizni kiriting!")
    application.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == '__main__':
    main()